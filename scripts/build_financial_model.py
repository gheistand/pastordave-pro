#!/usr/bin/env python3
"""
Pastor Dave Pro — Financial Model Spreadsheet Builder
Generates a multi-sheet Excel workbook with cost tables, revenue projections,
and charts showing break-even scenarios.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabel
import math

# ── COLORS ────────────────────────────────────────────────────────────────────
BRAND_BROWN  = "7C4F2A"
CREAM        = "F8F5F0"
WHITE        = "FFFFFF"
DARK_GRAY    = "2D2D2D"
MID_GRAY     = "6B6B6B"
LIGHT_GRAY   = "E5E0D8"
GREEN_PROFIT = "22C55E"
RED_LOSS     = "EF4444"
BLUE_PRO     = "3B82F6"
AMBER_START  = "F59E0B"
PURPLE_GROW  = "8B5CF6"
TEAL_COMBO   = "14B8A6"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=DARK_GRAY, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def border_thin():
    s = Side(style="thin", color=LIGHT_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="medium", color=BRAND_BROWN)
    return Border(bottom=s)

def money(ws, cell):
    ws[cell].number_format = '"$"#,##0.00'

def money_int(ws, cell):
    ws[cell].number_format = '"$"#,##0'

def pct(ws, cell):
    ws[cell].number_format = '0.0%'

def header_row(ws, row, cols, values, bg=BRAND_BROWN, fg=WHITE, size=11, bold=True):
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=bold, color=fg, size=size)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()

def data_cell(ws, row, col, value, bold=False, align="right", num_format=None, bg=WHITE, fg=DARK_GRAY):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border_thin()
    c.fill = fill(bg)
    if num_format:
        c.number_format = num_format
    return c

def section_header(ws, row, col, text, ncols=1, bg=LIGHT_GRAY):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=BRAND_BROWN, size=12)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_bottom()
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ncols-1)

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS / MODEL PARAMS
# ══════════════════════════════════════════════════════════════════════════════

CREDITS_PER_ACTIVE_USER = 71_000   # from real NHC data (30-day avg)

EL_PLANS = [
    {"name": "Pro",      "cost": 99,    "credits": 500_000},
    {"name": "Scale",    "cost": 330,   "credits": 2_000_000},
    {"name": "Business", "cost": 1_320, "credits": 11_000_000},
]

def el_cost_and_plan(active_users):
    needed = active_users * CREDITS_PER_ACTIVE_USER
    for p in EL_PLANS:
        if needed <= p["credits"]:
            return p["cost"], p["name"]
    overage = needed - 11_000_000
    return 1_320 + overage / 1000 * 0.12, "Business+"

FIXED_COSTS = {
    "API.Bible Pro (commercial)":     29,
    "API.Bible NLT Translation License": 10,
    "Cloudflare Workers Paid":         5,
    "Clerk Auth (free <50K users)":    0,
    "Mem0 Memory API (free tier)":     0,
    "Resend Email (free tier)":        0,
    "AssemblyAI Transcription (free)": 0,
}
FIXED_TOTAL = sum(FIXED_COSTS.values())  # $44

TIER_PRICES = {
    "Personal Pro":   20,
    "Church Starter": 149,
    "Church Growth":  249,
}

# Active users per subscriber (worst-case: all users are active)
ACTIVE_PER_TIER = {
    "Personal Pro":   1.0,    # 1 active per PP subscriber
    "Church Starter": 5.0,    # avg 5 active members per church starter
    "Church Growth":  15.0,   # avg 15 active members per church growth
}

def stripe_fee(revenue, n_transactions):
    return revenue * 0.029 + n_transactions * 0.30

def total_cost(pp, cs, cg):
    """Total monthly cost given subscriber counts."""
    active = pp * ACTIVE_PER_TIER["Personal Pro"] + \
             cs * ACTIVE_PER_TIER["Church Starter"] + \
             cg * ACTIVE_PER_TIER["Church Growth"]
    el, plan = el_cost_and_plan(active)
    rev = pp * TIER_PRICES["Personal Pro"] + \
          cs * TIER_PRICES["Church Starter"] + \
          cg * TIER_PRICES["Church Growth"]
    n_tx = pp + cs + cg
    stripe = stripe_fee(rev, n_tx)
    return FIXED_TOTAL + el + stripe, el, plan, stripe

def total_revenue(pp, cs, cg):
    return (pp * TIER_PRICES["Personal Pro"] +
            cs * TIER_PRICES["Church Starter"] +
            cg * TIER_PRICES["Church Growth"])

def profit(pp, cs, cg):
    rev = total_revenue(pp, cs, cg)
    cost, _, _, _ = total_cost(pp, cs, cg)
    return rev - cost


# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1: COST BREAKDOWN
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Cost Breakdown"
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 28
ws1.row_dimensions[1].height = 40

# Title
ws1.merge_cells("A1:D1")
c = ws1["A1"]
c.value = "Pastor Dave Pro — Monthly Cost Structure"
c.font = font(bold=True, color=WHITE, size=16)
c.fill = fill(BRAND_BROWN)
c.alignment = Alignment(horizontal="center", vertical="center")

# Fixed costs section
section_header(ws1, 3, 1, "FIXED MONTHLY COSTS (platform baseline)", 4)
header_row(ws1, 4, [1,2,3,4], ["Service", "Monthly Cost", "Plan/Tier", "Notes"], bg="4A3520")

row = 5
for name, cost in FIXED_COSTS.items():
    notes = {
        "API.Bible Pro (commercial)": "Required for commercial use",
        "API.Bible NLT Translation License": "NLT copyrighted — required for commercial use",
        "Cloudflare Workers Paid": "Pages Functions, D1 database, Workers",
        "Clerk Auth (free <50K users)": "Free up to 50,000 monthly retained users",
        "Mem0 Memory API (free tier)": "10K memories free; upgrade at scale",
        "Resend Email (free tier)": "3,000 emails/mo free; upgrade at scale",
        "AssemblyAI Transcription (free)": "5 hrs/mo free — covers 1 sermon/week",
    }.get(name, "")
    bg = CREAM if row % 2 == 0 else WHITE
    data_cell(ws1, row, 1, name, align="left", bg=bg)
    data_cell(ws1, row, 2, cost, num_format='"$"#,##0', bg=bg, bold=(cost > 0))
    data_cell(ws1, row, 3, "Free" if cost == 0 else "Paid", align="center", bg=bg,
              fg=GREEN_PROFIT if cost == 0 else DARK_GRAY)
    data_cell(ws1, row, 4, notes, align="left", bg=bg)
    row += 1

# Total fixed
data_cell(ws1, row, 1, "TOTAL FIXED COSTS", bold=True, align="left", bg=LIGHT_GRAY)
data_cell(ws1, row, 2, FIXED_TOTAL, bold=True, num_format='"$"#,##0', bg=LIGHT_GRAY, fg=BRAND_BROWN)
data_cell(ws1, row, 3, "", bg=LIGHT_GRAY)
data_cell(ws1, row, 4, "Before ElevenLabs (variable)", align="left", bg=LIGHT_GRAY)
row += 2

# ElevenLabs variable costs
section_header(ws1, row, 1, "ELEVENLABS VOICE AI — VARIABLE COST (scales with active users)", 4)
row += 1
header_row(ws1, row, [1,2,3,4],
    ["Active Users (all tiers combined)", "Credits Needed", "ElevenLabs Plan", "Monthly Cost"],
    bg="4A3520")
row += 1

user_counts = [1, 3, 5, 7, 8, 10, 12, 15, 20, 25, 28, 29, 30, 40, 50, 75, 100, 150]
prev_plan = None
for users in user_counts:
    cost, plan = el_cost_and_plan(users)
    credits = users * CREDITS_PER_ACTIVE_USER
    bg = "FEF9C3" if plan != prev_plan and prev_plan is not None else (CREAM if row % 2 == 0 else WHITE)
    note = " ← PLAN JUMP" if plan != prev_plan and prev_plan is not None else ""
    data_cell(ws1, row, 1, users, align="center", bg=bg)
    data_cell(ws1, row, 2, credits, num_format="#,##0", align="right", bg=bg)
    data_cell(ws1, row, 3, plan + note, align="center", bg=bg,
              fg=RED_LOSS if "JUMP" in note else DARK_GRAY,
              bold="JUMP" in note)
    data_cell(ws1, row, 4, cost, num_format='"$"#,##0', bg=bg,
              bold="JUMP" in note, fg=RED_LOSS if "JUMP" in note else DARK_GRAY)
    prev_plan = plan
    row += 1

row += 2
# Key insight box
section_header(ws1, row, 1, "⚠️  CRITICAL PRICING CLIFFS", 4, bg="FEF9C3")
row += 1
notes = [
    ("At 7 active users",  "ElevenLabs Pro $99/mo",   "Adding user 8 jumps cost to Scale $330 (+$231 overnight)"),
    ("At 28 active users", "ElevenLabs Scale $330/mo", "Adding user 29 jumps to Business $1,320 (+$990 overnight)"),
    ("Active user def.",   "3+ sessions/month",        "Members using it <3x/month don't count toward credit pool"),
]
for label, plan_name, detail in notes:
    bg = "FEF9C3"
    data_cell(ws1, row, 1, label, bold=True, align="left", bg=bg, fg=RED_LOSS)
    data_cell(ws1, row, 2, plan_name, align="center", bg=bg)
    data_cell(ws1, row, 3, detail, align="left", bg=bg)
    data_cell(ws1, row, 4, "", bg=bg)
    row += 1

# NHC baseline
row += 1
section_header(ws1, row, 1, "NHC PILOT — ACTUAL 30-DAY USAGE (Feb 26–Mar 26, 2026)", 4)
row += 1
nhc_data = [
    ("Total sessions", "218 (179 successful, 37 errors)", "", ""),
    ("Total voice minutes", "445 minutes (7.4 hours)", "", ""),
    ("Total credits used", "352,694", "", "From ElevenLabs usage report"),
    ("Credits per minute", "~792", "", ""),
    ("Average session length", "2.5 minutes", "", "Median: 1.8 min"),
    ("Active users (NHC)", "~5", "", "Testing/pilot group"),
    ("Cost per session", "~$0.39", "", "At Pro plan ($99/500K credits)"),
    ("Cost per active user/mo", "~$14", "", "71K credits × $99/500K"),
    ("Current monthly cost", "$143", "", "$99 ElevenLabs + $44 fixed"),
    ("Current revenue", "$0", "", "Free pilot — NHC not yet paying"),
    ("Monthly loss (pilot)", "-$143", "", "Investment in NHC relationship"),
]
header_row(ws1, row, [1,2,3,4], ["Metric", "Value", "", "Notes"], bg="4A3520")
row += 1
for i, (metric, val, _, note) in enumerate(nhc_data):
    bg = CREAM if i % 2 == 0 else WHITE
    data_cell(ws1, row, 1, metric, align="left", bg=bg)
    data_cell(ws1, row, 2, val, align="center", bg=bg, bold=True)
    data_cell(ws1, row, 3, "", bg=bg)
    data_cell(ws1, row, 4, note, align="left", bg=bg)
    row += 1


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2: REVENUE MODEL
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Revenue Model")
for col, width in zip("ABCDEFGHIJ", [6, 14, 16, 16, 14, 16, 16, 14, 14, 14]):
    ws2.column_dimensions[get_column_letter(ord(col)-ord('A')+1)].width = width

ws2.merge_cells("A1:J1")
c = ws2["A1"]
c.value = "Pastor Dave Pro — Revenue & Break-Even Model"
c.font = font(bold=True, color=WHITE, size=16)
c.fill = fill(BRAND_BROWN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

# Assumptions box
section_header(ws2, 3, 1, "MODEL ASSUMPTIONS", 10)
assumptions = [
    ("Personal Pro price", f"${TIER_PRICES['Personal Pro']}/mo", "Active users per PP subscriber", f"{ACTIVE_PER_TIER['Personal Pro']}"),
    ("Church Starter price", f"${TIER_PRICES['Church Starter']}/mo", "Avg active members per Church Starter", f"{ACTIVE_PER_TIER['Church Starter']}"),
    ("Church Growth price", f"${TIER_PRICES['Church Growth']}/mo", "Avg active members per Church Growth", f"{ACTIVE_PER_TIER['Church Growth']}"),
    ("Credits per active user/mo", f"{CREDITS_PER_ACTIVE_USER:,}", "Fixed costs", f"${FIXED_TOTAL}/mo"),
    ("Stripe fee", "2.9% + $0.30/tx", "ElevenLabs Pro ceiling", "7 active users"),
]
ws2.merge_cells("A4:E4")
ws2["A4"].value = "Input"
ws2["A4"].font = font(bold=True, color=BRAND_BROWN)
ws2.merge_cells("F4:J4")
ws2["F4"].value = "Input"
ws2["F4"].font = font(bold=True, color=BRAND_BROWN)

for i, (l1, v1, l2, v2) in enumerate(assumptions):
    r = 5 + i
    bg = CREAM if i % 2 == 0 else WHITE
    for col, val, bold in [(1,l1,False),(2,v1,True),(3,"",False),(4,"",False),(5,"",False),
                           (6,l2,False),(7,v2,True),(8,"",False),(9,"",False),(10,"",False)]:
        c = ws2.cell(row=r, column=col, value=val)
        c.font = font(bold=bold, color=BRAND_BROWN if bold else DARK_GRAY)
        c.fill = fill(bg)
        c.border = border_thin()
        c.alignment = Alignment(horizontal="left" if not bold else "center", vertical="center")

row = 12

# ── SCENARIO TABLE: All PP, no churches ──
section_header(ws2, row, 1, "SCENARIO A — Personal Pro Only", 10, bg="EFF6FF")
row += 1
header_row(ws2, row, range(1,11),
    ["PP Subs", "Revenue", "EL Cost", "EL Plan", "Fixed", "Stripe", "Total Cost", "Profit/Loss", "Margin %", "Break-Even?"],
    bg=BLUE_PRO)
row += 1
pp_data_rows = []
for pp in [1,2,3,5,7,8,10,12,15,20,25,28,29,30,35,40,50,75,100]:
    rev = total_revenue(pp, 0, 0)
    cost, el, plan, stripe = total_cost(pp, 0, 0)
    p = rev - cost
    margin = p / rev if rev > 0 else -1
    be = "✓ PROFIT" if p > 0 else ("≈ EVEN" if abs(p) < 15 else "LOSS")
    bg = "DCFCE7" if p > 0 else ("FEF9C3" if abs(p) < 15 else (WHITE if row%2==0 else CREAM))
    fg_p = GREEN_PROFIT if p > 0 else RED_LOSS
    vals = [pp, rev, el, plan, FIXED_TOTAL, round(stripe,2), round(cost,2), round(p,2), margin, be]
    fmts = ["#,##0", '"$"#,##0', '"$"#,##0', "@", '"$"#,##0', '"$"#,##0.00', '"$"#,##0', '"$"#,##0', "0.0%", "@"]
    for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.number_format = fmt
        c.fill = fill(bg)
        c.font = font(bold=(col == 8), color=fg_p if col == 8 else DARK_GRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()
    pp_data_rows.append(row)
    row += 1

row += 1

# ── SCENARIO TABLE: All Church Starter, no PP ──
section_header(ws2, row, 1, "SCENARIO B — Church Starter Only", 10, bg="FFFBEB")
row += 1
header_row(ws2, row, range(1,11),
    ["CS Churches", "Revenue", "EL Cost", "EL Plan", "Fixed", "Stripe", "Total Cost", "Profit/Loss", "Margin %", "Break-Even?"],
    bg=AMBER_START)
row += 1
for cs in [1,2,3,4,5,6,7,8,10,12,15,20,25,30,40,50]:
    rev = total_revenue(0, cs, 0)
    cost, el, plan, stripe = total_cost(0, cs, 0)
    p = rev - cost
    margin = p / rev if rev > 0 else -1
    be = "✓ PROFIT" if p > 0 else ("≈ EVEN" if abs(p) < 15 else "LOSS")
    bg = "DCFCE7" if p > 0 else ("FEF9C3" if abs(p) < 15 else (WHITE if row%2==0 else CREAM))
    vals = [cs, rev, el, plan, FIXED_TOTAL, round(stripe,2), round(cost,2), round(p,2), margin, be]
    fmts = ["#,##0", '"$"#,##0', '"$"#,##0', "@", '"$"#,##0', '"$"#,##0.00', '"$"#,##0', '"$"#,##0', "0.0%", "@"]
    for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.number_format = fmt
        c.fill = fill(bg)
        c.font = font(bold=(col == 8), color=(GREEN_PROFIT if p > 0 else RED_LOSS) if col == 8 else DARK_GRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()
    row += 1

row += 1

# ── SCENARIO TABLE: Mixed model (1:2 church:PP ratio) ──
section_header(ws2, row, 1, "SCENARIO C — Mixed Model (Churches + Personal Pro, 1:2 ratio)", 10, bg="F0FDF4")
row += 1
header_row(ws2, row, range(1,11),
    ["CS Churches", "PP Subs", "Total Revenue", "EL Cost", "EL Plan", "Fixed", "Stripe", "Total Cost", "Profit/Loss", "Break-Even?"],
    bg=TEAL_COMBO)
row += 1
for cs in [1,2,3,4,5,6,7,8,10,12,15,20,25,30,40,50]:
    pp = cs * 2
    rev = total_revenue(pp, cs, 0)
    cost, el, plan, stripe = total_cost(pp, cs, 0)
    p = rev - cost
    be = "✓ PROFIT" if p > 0 else ("≈ EVEN" if abs(p) < 15 else "LOSS")
    bg = "DCFCE7" if p > 0 else ("FEF9C3" if abs(p) < 20 else (WHITE if row%2==0 else CREAM))
    vals = [cs, pp, rev, el, plan, FIXED_TOTAL, round(stripe,2), round(cost,2), round(p,2), be]
    fmts = ["#,##0","#,##0",'"$"#,##0','"$"#,##0',"@",'"$"#,##0','"$"#,##0.00','"$"#,##0','"$"#,##0',"@"]
    for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.number_format = fmt
        c.fill = fill(bg)
        c.font = font(bold=(col == 9), color=(GREEN_PROFIT if p > 0 else RED_LOSS) if col == 9 else DARK_GRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()
    row += 1


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3: BREAK-EVEN MATRIX
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Break-Even Matrix")

ws3.merge_cells("A1:N1")
c = ws3["A1"]
c.value = "Profit/Loss Matrix — Church Starter × Personal Pro Combinations"
c.font = font(bold=True, color=WHITE, size=14)
c.fill = fill(BRAND_BROWN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

ws3.merge_cells("A2:N2")
ws3["A2"].value = "Green = Profitable | Yellow = Near Break-Even (±$50) | Red = Loss | Numbers = Monthly Profit/Loss $"
ws3["A2"].font = font(italic=True, color=MID_GRAY, size=10)
ws3["A2"].alignment = Alignment(horizontal="center")

# Column headers = PP subscribers
pp_vals = [0, 5, 10, 15, 20, 25, 30, 40, 50, 75, 100, 150, 200]
cs_vals = [0, 1, 2, 3, 4, 5, 6, 7, 8, 10, 12, 15, 20]

ws3.cell(row=4, column=1, value="CS →\nPP ↓").font = font(bold=True, color=BRAND_BROWN)
ws3.cell(row=4, column=1).alignment = Alignment(horizontal="center", wrap_text=True)

for j, pp in enumerate(pp_vals):
    c = ws3.cell(row=4, column=j+2, value=f"{pp} PP")
    c.font = font(bold=True, color=WHITE)
    c.fill = fill(BLUE_PRO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.column_dimensions[get_column_letter(j+2)].width = 11

ws3.column_dimensions["A"].width = 14

for i, cs in enumerate(cs_vals):
    row = i + 5
    ws3.row_dimensions[row].height = 20
    c = ws3.cell(row=row, column=1, value=f"{cs} Church")
    c.font = font(bold=True, color=WHITE)
    c.fill = fill(AMBER_START)
    c.alignment = Alignment(horizontal="center", vertical="center")

    for j, pp in enumerate(pp_vals):
        p = profit(pp, cs, 0)
        col = j + 2
        if p > 50:
            bg = "16A34A"
            fg = WHITE
        elif p > 0:
            bg = "86EFAC"
            fg = DARK_GRAY
        elif abs(p) <= 50:
            bg = "FDE68A"
            fg = DARK_GRAY
        elif p > -200:
            bg = "FCA5A5"
            fg = DARK_GRAY
        else:
            bg = "DC2626"
            fg = WHITE
        cell = ws3.cell(row=row, column=col, value=round(p))
        cell.number_format = '"$"#,##0;[RED]-"$"#,##0'
        cell.fill = fill(bg)
        cell.font = font(color=fg, size=9, bold=(abs(p) < 100))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin()

# Legend
ws3.cell(row=20, column=1, value="Legend:").font = font(bold=True, color=BRAND_BROWN)
legend = [
    ("16A34A", WHITE, "> $50 profit"),
    ("86EFAC", DARK_GRAY, "$0–$50 profit"),
    ("FDE68A", DARK_GRAY, "±$50 break-even"),
    ("FCA5A5", DARK_GRAY, "$0–$200 loss"),
    ("DC2626", WHITE, "> $200 loss"),
]
for k, (bg, fg, label) in enumerate(legend):
    c = ws3.cell(row=21+k, column=1, value=label)
    c.fill = fill(bg)
    c.font = font(color=fg, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_thin()


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4: GROWTH SCENARIOS
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Growth Scenarios")
ws4.column_dimensions["A"].width = 22
for col in "BCDEFGHIJ":
    ws4.column_dimensions[col].width = 14

ws4.merge_cells("A1:I1")
c = ws4["A1"]
c.value = "Pastor Dave Pro — 18-Month Growth Scenarios"
c.font = font(bold=True, color=WHITE, size=16)
c.fill = fill(BRAND_BROWN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 40

months = ["Mo 1","Mo 2","Mo 3","Mo 4","Mo 5","Mo 6","Mo 9","Mo 12","Mo 18"]

# Four growth scenarios
scenarios = {
    "Conservative\n(slow church adoption)": [
        # (PP, CS, CG)
        (0,0,0),(0,0,0),(1,0,0),(3,0,0),(5,1,0),(5,1,0),(8,2,0),(10,3,0),(15,5,0)
    ],
    "Moderate\n(1-2 churches/quarter)": [
        (0,0,0),(2,0,0),(5,1,0),(8,1,0),(10,2,0),(15,2,0),(20,4,0),(30,6,0),(50,10,0)
    ],
    "Optimistic\n(CTS partnership)": [
        (0,1,0),(5,2,0),(10,5,0),(15,8,0),(20,10,0),(30,15,0),(50,25,0),(75,40,0),(100,60,1)
    ],
    "Church Growth tier\n(premium upsell)": [
        (0,0,0),(0,1,0),(3,1,0),(5,2,0),(8,2,1),(10,3,1),(15,5,2),(25,8,3),(40,12,5)
    ],
}

scenario_colors = [BLUE_PRO, AMBER_START, GREEN_PROFIT, PURPLE_GROW]

# Data table
row = 3
section_header(ws4, row, 1, "Monthly Revenue & Profit by Growth Scenario", 9)
row += 1

header_row(ws4, row, range(1,10), ["Scenario"] + months, bg="4A3520")
row += 1

chart_data_start = row
chart_profit_data = {}

for s_idx, (scenario_name, month_subs) in enumerate(scenarios.items()):
    rev_row = row
    cost_row = row + 1
    profit_row = row + 2

    sc = scenario_colors[s_idx]
    # Revenue row
    c = ws4.cell(row=rev_row, column=1, value=f"{scenario_name}\nRevenue")
    c.font = font(bold=True, color=sc)
    c.fill = fill(CREAM)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = border_thin()
    ws4.row_dimensions[rev_row].height = 30

    # Cost row
    c = ws4.cell(row=cost_row, column=1, value="  Cost")
    c.font = font(italic=True, color=MID_GRAY)
    c.fill = fill(CREAM)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_thin()

    # Profit row
    c = ws4.cell(row=profit_row, column=1, value="  Net Profit")
    c.font = font(bold=True)
    c.fill = fill(CREAM)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_thin()

    profit_vals = []
    for col_idx, (pp, cs, cg) in enumerate(month_subs):
        col = col_idx + 2
        rev = total_revenue(pp, cs, cg)
        cost, el, plan, stripe = total_cost(pp, cs, cg)
        p = rev - cost

        for r, val in [(rev_row, rev), (cost_row, cost), (profit_row, p)]:
            is_profit = r == profit_row
            c = ws4.cell(row=r, column=col, value=round(val))
            c.number_format = '"$"#,##0'
            bg = CREAM
            if is_profit:
                bg = "DCFCE7" if val > 0 else ("FEF9C3" if abs(val) < 50 else "FEE2E2")
                c.font = font(bold=True, color=GREEN_PROFIT if val > 0 else RED_LOSS)
            else:
                c.font = font(color=DARK_GRAY)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border_thin()

        profit_vals.append(round(p))

    chart_profit_data[scenario_name.replace("\n", " ")] = (profit_row, profit_vals)
    row += 4  # 3 rows + 1 blank


# ── CHART 1: Revenue per scenario line chart ──────────────────────────────────
ws5 = wb.create_sheet("Charts")
ws5.merge_cells("A1:P1")
c = ws5["A1"]
c.value = "Pastor Dave Pro — Visual Break-Even Analysis"
c.font = font(bold=True, color=WHITE, size=16)
c.fill = fill(BRAND_BROWN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 40

# Build chart data on ws5 directly (avoids cross-sheet chart issues)
ws5.cell(row=3, column=1, value="Month").font = font(bold=True, color=BRAND_BROWN)
month_labels = ["1","2","3","4","5","6","9","12","18"]
for j, m in enumerate(month_labels):
    c = ws5.cell(row=3, column=j+2, value=int(m))
    c.font = font(bold=True)

scenario_rows = {}
for i, (scenario_name, month_subs) in enumerate(scenarios.items()):
    row = 4 + i
    short_name = {
        "Conservative\n(slow church adoption)": "Conservative",
        "Moderate\n(1-2 churches/quarter)": "Moderate",
        "Optimistic\n(CTS partnership)": "Optimistic (CTS)",
        "Church Growth tier\n(premium upsell)": "With Church Growth"
    }[scenario_name]
    ws5.cell(row=row, column=1, value=short_name).font = font(bold=True)
    for j, (pp, cs, cg) in enumerate(month_subs):
        p = profit(pp, cs, cg)
        c = ws5.cell(row=row, column=j+2, value=round(p))
        c.number_format = '"$"#,##0'
    scenario_rows[short_name] = row

# Chart 1: Profit/Loss by scenario
chart1 = LineChart()
chart1.title = "Monthly Net Profit by Growth Scenario"
chart1.style = 10
chart1.y_axis.title = "Net Profit / Loss ($)"
chart1.x_axis.title = "Month"
chart1.width = 22
chart1.height = 14
chart1.y_axis.crossAx = 500
chart1.x_axis.crossAx = 500

colors_chart = ["4472C4", "ED7D31", "70AD47", "7030A0"]
for i, (name, crow) in enumerate(scenario_rows.items()):
    data = Reference(ws5, min_col=2, max_col=10, min_row=crow, max_row=crow)
    chart1.add_data(data, titles_from_data=False)
    chart1.series[i].title = SeriesLabel(v=name)
    chart1.series[i].graphicalProperties.line.solidFill = colors_chart[i]
    chart1.series[i].graphicalProperties.line.width = 25000

cats = Reference(ws5, min_col=2, max_col=10, min_row=3)
chart1.set_categories(cats)

# Add zero line reference
zero_row = 8
for j in range(1, 11):
    ws5.cell(row=zero_row, column=j, value=0 if j > 1 else "Break-Even ($0)")
ws5.cell(row=zero_row, column=1).font = font(italic=True, color=RED_LOSS)
zero_ref = Reference(ws5, min_col=2, max_col=10, min_row=zero_row)
chart1.add_data(zero_ref, titles_from_data=False)
chart1.series[-1].title = SeriesLabel(v="Break-Even ($0)")
chart1.series[-1].graphicalProperties.line.solidFill = "FF0000"
chart1.series[-1].graphicalProperties.line.dashDot = "dash"
chart1.series[-1].graphicalProperties.line.width = 15000

ws5.add_chart(chart1, "A10")


# Chart 2: Bar chart — revenue components at key milestones
# Build data: 5 milestone scenarios
ws5.cell(row=35, column=1, value="Scenario").font = font(bold=True, color=BRAND_BROWN)
ws5.cell(row=35, column=2, value="PP Revenue").font = font(bold=True, color=BLUE_PRO)
ws5.cell(row=35, column=3, value="CS Revenue").font = font(bold=True, color=AMBER_START)
ws5.cell(row=35, column=4, value="CG Revenue").font = font(bold=True, color=PURPLE_GROW)
ws5.cell(row=35, column=5, value="Total Cost").font = font(bold=True, color=RED_LOSS)
ws5.cell(row=35, column=6, value="Net Profit").font = font(bold=True, color=GREEN_PROFIT)

milestones = [
    ("NHC Pilot (free)", 0, 0, 0),
    ("1 Church + 5 PP", 5, 1, 0),
    ("3 Churches + 15 PP", 15, 3, 0),
    ("5 Churches + 25 PP", 25, 5, 0),
    ("10 Churches + 50 PP", 50, 10, 0),
    ("15 Churches + 75 PP", 75, 15, 0),
    ("20 Churches + 100 PP", 100, 20, 0),
    ("30 Ch + 5 Growth + 100 PP", 100, 20, 5),
    ("50 Ch + 10 Growth + 200 PP", 200, 40, 10),
]

for i, (name, pp, cs, cg) in enumerate(milestones):
    row = 36 + i
    rev_pp = pp * TIER_PRICES["Personal Pro"]
    rev_cs = cs * TIER_PRICES["Church Starter"]
    rev_cg = cg * TIER_PRICES["Church Growth"]
    cost, el, plan, stripe = total_cost(pp, cs, cg)
    p = (rev_pp + rev_cs + rev_cg) - cost
    ws5.cell(row=row, column=1, value=name).font = font(size=9)
    ws5.cell(row=row, column=2, value=rev_pp).number_format = '"$"#,##0'
    ws5.cell(row=row, column=3, value=rev_cs).number_format = '"$"#,##0'
    ws5.cell(row=row, column=4, value=rev_cg).number_format = '"$"#,##0'
    ws5.cell(row=row, column=5, value=round(cost)).number_format = '"$"#,##0'
    ws5.cell(row=row, column=6, value=round(p)).number_format = '"$"#,##0'
    ws5.cell(row=row, column=6).font = font(bold=True, color=GREEN_PROFIT if p > 0 else RED_LOSS)

chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "Revenue Components vs. Cost — Key Milestones"
chart2.y_axis.title = "Monthly $ Amount"
chart2.x_axis.title = "Subscriber Scenario"
chart2.width = 24
chart2.height = 16

for col_idx, (label, color) in enumerate([
    ("PP Revenue", "4472C4"),
    ("CS Revenue", "ED7D31"),
    ("CG Revenue", "7030A0"),
    ("Total Cost", "FF0000"),
], start=2):
    data = Reference(ws5, min_col=col_idx, max_col=col_idx, min_row=35, max_row=35+len(milestones))
    chart2.add_data(data, titles_from_data=True)
    chart2.series[-1].graphicalProperties.solidFill = color

chart2.set_categories(Reference(ws5, min_col=1, max_col=1, min_row=36, max_row=35+len(milestones)))
ws5.add_chart(chart2, "A55")


# Chart 3: PP-only break-even
ws5.cell(row=90, column=1, value="PP Subs").font = font(bold=True)
ws5.cell(row=90, column=2, value="Revenue").font = font(bold=True, color=BLUE_PRO)
ws5.cell(row=90, column=3, value="Cost").font = font(bold=True, color=RED_LOSS)
ws5.cell(row=90, column=4, value="Profit").font = font(bold=True, color=GREEN_PROFIT)

pp_range = [1,3,5,7,8,10,12,15,20,25,28,29,30,35,40,50,75,100]
for i, pp in enumerate(pp_range):
    row = 91 + i
    rev = total_revenue(pp, 0, 0)
    cost, _, _, _ = total_cost(pp, 0, 0)
    p = rev - cost
    ws5.cell(row=row, column=1, value=pp)
    ws5.cell(row=row, column=2, value=rev).number_format = '"$"#,##0'
    ws5.cell(row=row, column=3, value=round(cost)).number_format = '"$"#,##0'
    ws5.cell(row=row, column=4, value=round(p)).number_format = '"$"#,##0'

chart3 = LineChart()
chart3.title = "Personal Pro Only — Revenue vs. Cost (Break-Even Analysis)"
chart3.style = 10
chart3.y_axis.title = "Monthly $ Amount"
chart3.x_axis.title = "Number of PP Subscribers"
chart3.width = 22
chart3.height = 14

rev_data = Reference(ws5, min_col=2, max_col=3, min_row=90, max_row=90+len(pp_range))
chart3.add_data(rev_data, titles_from_data=True)
chart3.series[0].graphicalProperties.line.solidFill = "4472C4"
chart3.series[1].graphicalProperties.line.solidFill = "FF0000"
chart3.series[0].graphicalProperties.line.width = 25000
chart3.series[1].graphicalProperties.line.width = 25000

chart3.set_categories(Reference(ws5, min_col=1, max_col=1, min_row=91, max_row=90+len(pp_range)))
ws5.add_chart(chart3, "M10")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5: PROFITABILITY SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Executive Summary")
ws6.column_dimensions["A"].width = 32
ws6.column_dimensions["B"].width = 22
ws6.column_dimensions["C"].width = 32

ws6.merge_cells("A1:C1")
c = ws6["A1"]
c.value = "Pastor Dave Pro — Executive Profitability Summary"
c.font = font(bold=True, color=WHITE, size=16)
c.fill = fill(BRAND_BROWN)
c.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 40

summaries = [
    ("BOTTOM LINE", None, None),
    ("Monthly platform cost (NHC pilot)", "$143/mo", "ElevenLabs $99 + Fixed $44"),
    ("Current revenue", "$0", "Free pilot — NHC not yet paying"),
    ("Monthly loss today", "-$143/mo", "Investment in NHC relationship"),
    ("", None, None),
    ("BREAK-EVEN THRESHOLDS", None, None),
    ("Fastest break-even", "1 church + 5 Personal Pro", "Revenue $199 > Cost ~$152"),
    ("Personal Pro only break-even", "28 subscribers", "Revenue $420 > Cost ~$395"),
    ("Church Starter only break-even", "5 churches", "Revenue $495 > Cost ~$427"),
    ("", None, None),
    ("PROFITABILITY MILESTONES", None, None),
    ("Meaningful profit (>$200/mo)", "5 churches + 25 PP", "~$440/mo net"),
    ("Small business viable (>$1K/mo)", "15 churches + 75 PP", "~$1,100/mo net"),
    ("Full-time income potential", "50 churches + 200 PP", "~$9,000/mo net"),
    ("CTS partnership (50 churches)", "Instant Scale", "~$5,000-10,000/mo net"),
    ("", None, None),
    ("KEY RISKS", None, None),
    ("ElevenLabs cliff at 8 active users", "Cost jumps $231/mo", "Plan Pro→Scale on user 8"),
    ("ElevenLabs cliff at 29 active users", "Cost jumps $990/mo", "Plan Scale→Business on user 29"),
    ("API.Bible commercial license", "$39/mo required", "Must upgrade before launch"),
    ("", None, None),
    ("RECOMMENDED ACTIONS", None, None),
    ("1. Charge NHC now (pilot rate)", "$49/mo pilot", "Offsets cost, establishes payment norm"),
    ("2. Launch Personal Pro immediately", "$20/mo", "Fastest path to break-even"),
    ("3. Upgrade API.Bible", "$39/mo", "Required for commercial launch"),
    ("4. Enforce active user cap on Starter", "7 active users max", "Prevents unexpected EL plan jump"),
    ("5. Pursue CTS partnership actively", "50+ church distribution", "Skips all break-even math"),
]

row = 3
for item in summaries:
    label, val, note = item
    if val is None:  # Section header
        ws6.row_dimensions[row].height = 28
        section_header(ws6, row, 1, label, 3)
    else:
        ws6.row_dimensions[row].height = 20
        bg = CREAM if row % 2 == 0 else WHITE
        c1 = ws6.cell(row=row, column=1, value=label)
        c1.font = font(color=DARK_GRAY)
        c1.fill = fill(bg)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.border = border_thin()
        c2 = ws6.cell(row=row, column=2, value=val)
        c2.font = font(bold=True, color=BRAND_BROWN)
        c2.fill = fill(bg)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = border_thin()
        c3 = ws6.cell(row=row, column=3, value=note)
        c3.font = font(italic=True, color=MID_GRAY)
        c3.fill = fill(bg)
        c3.alignment = Alignment(horizontal="left", vertical="center")
        c3.border = border_thin()
    row += 1


# ── SAVE ──────────────────────────────────────────────────────────────────────
output_path = "/Users/glennheistand/Projects/pastordave-pro/PastorDavePro_Financial_Model.xlsx"
wb.save(output_path)
print(f"✅ Saved: {output_path}")
print(f"   Sheets: {[s.title for s in wb.worksheets]}")